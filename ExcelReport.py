"""
Created on 2020/05/23

@author: kevin
"""
#!/usr/bin/python
import openpyxl
from openpyxl import Workbook  
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment,  numbers
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook

import logging, time, configparser, os

# Setup logger
logger = logging.getLogger(__name__)

def write_to_excel(report_name, column, data, file_name):
    
    try:
        logger.info(f"start generating excel file {file_name}..")

        wb = Workbook()  
        ws = wb.active  

        ws.merge_cells('a1:y1')
        ws['A1'] = f"{report_name}"
        cell = ws['A1']
        cell.font = Font(size =20, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        col_l = []
        for col in column:
            col_l.append(col)

        ws.append(col_l)

        row_count = 0
        for rec in data:
            ws.append(list(rec))
            row_count = row_count + 1
        
        wb.save(f"{file_name}")  
        logger.info(f"Total {row_count} records inserted into excel file")

        # doing formatting of the excel
        wb = load_workbook(filename = f"{file_name}") 

        ws = wb["Sheet"]
        
        #for col in range(1, 26):
        for col in range(len(col_l)):            
            cell = ws.cell(column=col+1, row=2) 
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col+1)].auto_size = True        

        for row in range(2, row_count+3):
            cell=ws.cell(column=1, row=row)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell=ws.cell(column=2, row=row)
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            
            cell=ws.cell(column=3, row=row)
            cell.number_format = u'#,##0.0000;'

        wb.save(f"{file_name}")  

        logger.info(f"Excel file has been formatted")
    
    except (Exception) as error:
        logger.warning("Something wrong in appending value to excel:")
        logger.exception(error)
        print(error)


if __name__ == '__main__':

    ExcelFileName = "test.xlsx"
    ExcelFileName_FullPath = os.path.join(os.path.dirname(os.path.realpath(__file__)), ExcelFileName)

    #column = [('title', 1043), ('number', 25), ('description', 25)]
    column = ['title', 'number', 'description']

    #data = [["col1 header", "col2 header","col3 header"],
    data = [["col1 row2", 12345.67, 0.1234567],
            ["col1 row3", 12345.67, 0.1234567]]          

    write_to_excel(report_name='test', column=column, data=data,  file_name= ExcelFileName_FullPath)

    wb = load_workbook(filename = f"{ExcelFileName_FullPath}") 
    ws = wb["Sheet"]
    
    print (ws.cell(2, 1).value)
    print (ws.cell(3, 1).value)

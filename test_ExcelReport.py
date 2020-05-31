import ExcelReport
import os
import pytest
from openpyxl.reader.excel import load_workbook

def test_Excel_report():

    ExcelFileName = "test.xlsx"
    ExcelFileName_FullPath = os.path.join(os.path.dirname(os.path.realpath(__file__)), ExcelFileName)
    os.remove(ExcelFileName_FullPath) if os.path.exists(ExcelFileName_FullPath) else None

    column = ['title', 'number', 'description']
    data = [["col1 row2", 12345.67, 0.1234567],
            ["col1 row3", 12345.67, 0.1234567]]          

    ExcelReport.write_to_excel(report_name='test', column=column, data=data,  file_name= ExcelFileName_FullPath)

    wb = load_workbook(filename = f"{ExcelFileName_FullPath}") 
    ws = wb["Sheet"]
    
    assert (ws.cell(2, 1).value, ws.cell(3, 1).value)  == ('title', 'col1 row2')
    
    

